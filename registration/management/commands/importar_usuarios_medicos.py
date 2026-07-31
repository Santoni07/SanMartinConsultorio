from openpyxl import load_workbook

from django.core.management.base import BaseCommand
from django.contrib.auth.models import User

from medicos.models import Medico


class Command(BaseCommand):

    help = "Importa usuarios para médicos desde un archivo Excel"

    def add_arguments(self, parser):

        parser.add_argument(
            "archivo",
            type=str,
            help="Ruta del archivo Excel"
        )

    def handle(self, *args, **options):

        archivo = options["archivo"]

        workbook = load_workbook(archivo)

        hoja = workbook.active

        creados = 0
        existentes = 0
        errores = 0

        # ===========================================
        # COLUMNAS
        # A = Nombre
        # B = Matrícula
        # C = Clave
        # D = Usuario
        # ===========================================
        for i in range(1, 10):
            print(i, [c.value for c in hoja[i]])
        
        
        for fila in hoja.iter_rows(min_row=2):
            print([c.value for c in fila])

            break

            nombre = str(fila[0].value).strip() if fila[0].value else ""

            matricula = str(fila[1].value).strip() if fila[1].value else ""

            clave = str(fila[2].value).strip() if fila[2].value else ""

            usuario = str(fila[3].value).strip() if fila[3].value else ""

            # ===========================================
            # VALIDACIONES
            # ===========================================

            if not matricula:

                self.stdout.write(
                    self.style.ERROR(
                        f"{nombre}: matrícula vacía."
                    )
                )

                errores += 1
                continue

            if not usuario:

                self.stdout.write(
                    self.style.ERROR(
                        f"{nombre}: usuario vacío."
                    )
                )

                errores += 1
                continue

            if not clave:

                self.stdout.write(
                    self.style.ERROR(
                        f"{nombre}: clave vacía."
                    )
                )

                errores += 1
                continue

            # ===========================================
            # BUSCAR MÉDICO
            # ===========================================

            try:

                medico = Medico.objects.get(
                    matricula=matricula
                )

            except Medico.DoesNotExist:

                self.stdout.write(
                    self.style.ERROR(
                        f"No existe un médico con matrícula {matricula}"
                    )
                )

                errores += 1
                continue

            # ===========================================
            # ¿YA TIENE USUARIO?
            # ===========================================

            if medico.user:

                self.stdout.write(
                    self.style.WARNING(
                        f"{medico.nombre} {medico.apellido} ya tiene un usuario asociado."
                    )
                )

                existentes += 1
                continue

            # ===========================================
            # ¿USERNAME EXISTENTE?
            # ===========================================

            if User.objects.filter(
                username=usuario
            ).exists():

                self.stdout.write(
                    self.style.WARNING(
                        f"El usuario '{usuario}' ya existe."
                    )
                )

                existentes += 1
                continue

            # ===========================================
            # CREAR USUARIO
            # ===========================================

            user = User.objects.create_user(

                username=usuario,

                email=medico.email,

                first_name=medico.nombre,

                last_name=medico.apellido,

                password=clave

            )

            # ===========================================
            # ASOCIAR AL MÉDICO
            # ===========================================

            medico.user = user

            medico.save()

            creados += 1

            self.stdout.write(

                self.style.SUCCESS(

                    f"✔ {medico.nombre} {medico.apellido} -> Usuario creado correctamente."

                )

            )

        # ===========================================
        # RESUMEN
        # ===========================================

        self.stdout.write("")
        self.stdout.write("=" * 60)

        self.stdout.write(
            self.style.SUCCESS(
                f"Usuarios creados : {creados}"
            )
        )

        self.stdout.write(
            self.style.WARNING(
                f"Ya existentes    : {existentes}"
            )
        )

        self.stdout.write(
            self.style.ERROR(
                f"Errores          : {errores}"
            )
        )

        self.stdout.write("=" * 60)