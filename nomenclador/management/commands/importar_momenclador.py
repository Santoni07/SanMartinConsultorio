from openpyxl import load_workbook

from django.core.management.base import BaseCommand

from nomenclador.models import NomencladorGeneral


class Command(BaseCommand):

    help = "Importa el Nomenclador General desde un archivo Excel."

    def add_arguments(self, parser):

        parser.add_argument(
            "archivo",
            type=str,
            help="Ruta del archivo Excel a importar."
        )

    def handle(self, *args, **options):

        archivo = options["archivo"]

        self.stdout.write(f"\nImportando archivo: {archivo}\n")

        wb = load_workbook(archivo)
        ws = wb.active

        creados = 0
        actualizados = 0
        omitidos = 0

        for fila in ws.iter_rows(min_row=2, values_only=True):

            # Ignorar filas completamente vacías
            if not fila or not fila[0]:
                omitidos += 1
                continue

            codigo = str(fila[0]).strip()

            descripcion = ""

            if len(fila) > 1 and fila[1]:
                descripcion = str(fila[1]).strip()

            _, created = NomencladorGeneral.objects.update_or_create(

                codigo=codigo,

                defaults={
                    "descripcion": descripcion,
                    "activo": True,
                }

            )

            if created:
                creados += 1
            else:
                actualizados += 1

        self.stdout.write(self.style.SUCCESS("\n======================================"))
        self.stdout.write(self.style.SUCCESS(" IMPORTACIÓN FINALIZADA "))
        self.stdout.write(self.style.SUCCESS("======================================"))
        self.stdout.write(self.style.SUCCESS(f"Prestaciones creadas.....: {creados}"))
        self.stdout.write(self.style.SUCCESS(f"Prestaciones actualizadas: {actualizados}"))
        self.stdout.write(self.style.SUCCESS(f"Filas omitidas...........: {omitidos}"))
        self.stdout.write(self.style.SUCCESS("======================================\n"))